from datetime import datetime, timedelta
import locale
import calendar
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.chart import BarChart, StockChart, LineChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.updown_bars import UpDownBars
from openpyxl.chart.data_source import NumData, NumVal
from openpyxl.chart.trendline import Trendline
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import NamedStyle, Font
from criar_excel import aplicar_estilo_area

locale.setlocale(locale.LC_ALL, ("pt_BR", "utf-8"))

def estilo_tabelas():
    """Cria o estilo utilizado para esconder as tabelas

    Returns:
        estilo (openpyxl.style): deve estar no formato  estilo_titulo_carteira = NamedStyle(name = "estilo_titulo_carteira")
    """    
    estilo = NamedStyle(name = "estilo")
    estilo.font = Font(color = "00FFFFFF")
    return estilo

def graf_barras(base, arquivo, estilo):
    """Gera um gráfico de barras em uma planilha excel já existente. É necessário que exista uma Worksheet na planilha chamada Estatísticas.

    Args:
        base (dict): Recebe um dict com o nome de uma ou mais ações como keys, não importando os valores relacionados
        arquivo (str): Nome do arquivo excel onde será gerado o gráfico. Deve estar no formato ("nome.xlsx")
        estilo (openpyxl.style): deve estar no formato  estilo_titulo_carteira = NamedStyle(name = "estilo_titulo_carteira")
    """    

    #O gráfico de barras mostra quanto uma ação subiu ou desceu de valor no último ano, e para isso devemos primeiro adquirir os dados pelo yfinance, e em seguida montar o gráfico
    
    #Primeiro, o código para pegar a data de um ano atrás
    dias = 365
    data_inicio = datetime.now() - timedelta(days=dias)
    
    #Vamos tratar o caso de termos um ano bissexto, por que não?
    ano = int(data_inicio.strftime("%Y"))
    if calendar.isleap(ano) == True:
        dias += 1
        data_inicio = datetime.now() - timedelta(days=dias)
    
    data_final = data_inicio + timedelta(days=1)

    #Também devemos formatar a string na forma que o Yfinance entende
    data_inicio = data_inicio.strftime("%Y-%m-%d")
    data_final = data_final.strftime("%Y-%m-%d")
    
    #Agora, vamos pegar os valores de cada ativo um ano atrás
    valores_antigos = {}
    for ativo in base.keys():
        ticket = yf.Ticker(ativo)
        ticket_hist = ticket.history(start=data_inicio, end=data_final, debug = False)
        ticket_hist.fillna(-1, inplace = True)

        # É muito comum que em determinado dia, a bolsa não funcione por vários motivos, e por isso o Yfinance 
        # não terá dados referentes àquele dia específico, e nos entregará um dataframe vazio, por isso 
        # devemos tratar este caso. Caso o dataframe retornado seja vazio, iremos pegar os dados de um dia 
        # antes, repetidamente, até ser retornado um dataframe válido

        # Também foi necessário tratar um caso em que o dataframe é retornado com alguns valores faltantes,
        # com as células contendo NaN. Para isso foi usado o método fillna do dataframe, e um Or no While

        # Caso após 15 dias não seja encontrado nenhum valor, o código irá desistir e partir para o próximo
        dias_antigos = dias
        falha = False
        while ticket_hist.size == 0 or ticket_hist.Close[0] == -1:

            if dias_antigos == 380:
                print(f"Não foi possível encontrar valores para {ativo}")
                falha = True
                break

            dias_antigos += 1
            inicio_altern = datetime.now() - timedelta(days=dias_antigos)
            final_altern = inicio_altern + timedelta(days=1)
            ticket_hist = ticket.history(start=inicio_altern, end=final_altern, debug = False)
            
            ticket_hist.fillna(-1, inplace = True)

        if falha == False:
            valores_antigos[ativo] = ticket_hist.Close[0]

    #Agora, fazemos o mesmo processo para os dados de cada ativo no dia atual
    valores_hoje = {}
    for ativo in base.keys():
        ticket = yf.Ticker(ativo)
        ticket_hist = ticket.history(period="1d", debug = False)
        ticket_hist.fillna(-1, inplace = True)

        antes_ontem = 1
        falha = False
        while ticket_hist.size == 0 or ticket_hist.Close[0] == -1 :
            
            if antes_ontem == 15: #Este if impede que o while fique em loop para sempre
                print(f"Não foi possível encontrar valores para {ativo}")
                falha = True
                break
            
            antes_ontem += 1
            inicio_altern = datetime.now() - timedelta(days=antes_ontem)
            final_altern = inicio_altern + timedelta(days=1)
            ticket_hist = ticket.history(start=inicio_altern, end=final_altern, debug = False)
            ticket_hist.fillna(-1, inplace = True)

        if falha == False:
            valores_hoje[ativo] = ticket_hist.Close[0]
    
    #Agora, vamos fazer a matemágica para chegar nos dados necessários
    dados = []
    for chave in valores_antigos.keys():
        resultado = (valores_hoje[chave] / valores_antigos[chave]) - 1
        dados.append(resultado)

    #E as categorias sendo as chaves
    categorias = []
    for chave in valores_antigos.keys():
        categorias.append(chave)

    #Agora que temos os dados, entramos na etapa de montar o gráfico:

    #Carregamos a planilha e selecionamos a folha de trabalho
    planilha = load_workbook(arquivo)
    estatisticas = planilha["Estatísticas"]

    #Inserimos os dados na planilha
    linha_d = 2
    for dado in dados:
        linha_d += 1
        estatisticas[f"C{linha_d}"].value = dado  
    
    linha_c = 2
    for cat in categorias:
        linha_c += 1
        estatisticas[f"D{linha_c}"].value = cat
    
    #Esconde a planilha
    aplicar_estilo_area(estatisticas,2,linha_d,3,4,estilo)

    #Selecionamos o tipo e os títulos do grafico
    bar_grafic = BarChart()
    bar_grafic.title = "Variação das Ações no Último Ano"
    bar_grafic.type = "col"
    bar_grafic.y_axis.title = "Variação"
    bar_grafic.y_axis.number_format = "0%"
    bar_grafic.x_axis.title = "Ativos"
    bar_grafic.x_axis.tickLblPos = "low"

    #Seleciona os dados do grafico
    dados_finais = Reference(estatisticas, min_col=3, min_row=3, max_col=3, max_row=f"{linha_d}")
    categorias_finais = Reference(estatisticas, min_col=4, min_row=3, max_col=4, max_row=f"{linha_c}")
    bar_grafic.add_data(dados_finais, titles_from_data=False)
    bar_grafic.set_categories(categorias_finais)

    #Modificando as cores das barras que representam valores menores que zero (Isso cria uma série nova pra cada barra negativa, então tecnicamente é uma gambiarra)
    series = bar_grafic.series[0]
    for i in range(len(dados)):
        if dados[i] < 0:
            pt = DataPoint(idx=i)
            pt.graphicalProperties.line.solidFill = "DC143C"
            pt.graphicalProperties.solidFill = "DC143C"
            series.dPt.append(pt)
        else:
            pt = DataPoint(idx=i)
            pt.graphicalProperties.line.solidFill = "00A300"
            pt.graphicalProperties.solidFill = "00A300"
            series.dPt.append(pt)
            
    #Estilizando
    bar_grafic.legend = None
    bar_grafic.style = 9
    bar_grafic.x_axis.graphicalProperties = GraphicalProperties(ln=LineProperties(w=3))
    bar_grafic.y_axis.majorGridlines.graphicalProperties = GraphicalProperties(ln=LineProperties(prstDash="dash"))

    #Adiciona o gráfico à planilha
    estatisticas.add_chart(bar_grafic, "B3")
    planilha.save(arquivo)

def graf_linha(quantias, base, arquivo, estilo):
    """Gera um gráfico de linhas em uma planilha excel já existente. É necessário que exista uma Worksheet na planilha chamada Estatísticas.

    Args:
        quantias (dict): Recebe um dicionário com os nomes das ações como chaves, e as quantias das ações na carteira como valores.
        base (dict): Recebe um dicionário com os nomes das ações como chaves (é importante que sejam iguais aos das quantias), e como valores, dataframes com os históricos de cada ação, no último ano.
        arquivo (str): Nome do arquivo excel onde será gerado o gráfico. Deve estar no formato ("nome.xlsx")
        estilo (openpyxl.style): deve estar no formato  estilo_titulo_carteira = NamedStyle(name = "estilo_titulo_carteira")
    """

    # Vamos começar importando a planilha
    planilha = load_workbook(arquivo)
    estatisticas = planilha["Estatísticas"]

    # Precisamos pegar ambos os dicionarios dados como argumento, e converter em um dicionário
    # para inserir na tabela.
    valor_por_ativo = {}
    for ativo in base.keys():
        dataframe = base[ativo]
        valor_por_ativo[ativo] = {}

        for data in dataframe.index:
            strf_data = data.strftime("%Y-%m-%d")
            valor = (dataframe.at[data, "Close"]) * float(quantias[ativo])
            valor_por_ativo[ativo][strf_data] = valor
        # Caso algum ativo possua menos de 20 dados no último ano, vamos ignorá-lo para não prejudicar
        # o resultado do gráfico, afinal ele afetará pouco o valor total da carteira
        if len(valor_por_ativo[ativo]) < 20:
            valor_por_ativo.pop(ativo)

    # Agora precisamos saber quais datas esses dicionários possuem em comum
    chave0 = list(valor_por_ativo.keys())[0]
    datas_comum = set(valor_por_ativo[chave0].keys())
    for ativo in valor_por_ativo.keys():
        datas_comum = datas_comum.intersection(set(valor_por_ativo[ativo].keys()))
    
    # Com as datas comuns em mãos, vamos unir o valor de todos os ativos em um só,
    # tendo assim o valor total da carteira por data.
    dados_finais = {}
    for data in valor_por_ativo[chave0].keys():
        valor_total = 0
        if data in datas_comum:
            for ativo in valor_por_ativo.keys():
                valor_total += valor_por_ativo[ativo][data]
            dados_finais[data] = valor_total

    # Inserindo os dados na planilha
    linha = 2
    coluna_data = 6
    coluna_valor = 7
    estatisticas.cell(row=linha, column=coluna_data, value="Data")
    estatisticas.cell(row=linha, column=coluna_valor, value="Valores")
    for key in dados_finais.keys():
        linha += 1
        estatisticas.cell(row=linha, column=coluna_data, value=key)
        estatisticas.cell(row=linha, column=coluna_valor, value=dados_finais[key])
    #Esconde a planilha
    aplicar_estilo_area(estatisticas,2,linha,coluna_data,coluna_valor,estilo)

    #Seleciona o tipo e os titulos
    graf_linha = LineChart()
    graf_linha.title = "Valor Total da Carteira ao Longo do Tempo"
    graf_linha.y_axis.title = "Valor da Carteira (em R$)"
    graf_linha.x_axis.title = "Data"

    #Seleciona os dados do grafico na planilha
    dados = Reference(estatisticas, min_col=7, min_row=2, max_col=7, max_row=linha)
    tempo = Reference(estatisticas, min_col=6, min_row=3, max_col=6, max_row=linha)
    graf_linha.add_data(dados, titles_from_data=True)
    graf_linha.set_categories(tempo)

    #Estilizando
    s1 = graf_linha.series[0]
    s1.graphicalProperties.line.solidFill = "0000FF"
    s1.graphicalProperties.line.width = 25000
    graf_linha.style = 13
    graf_linha.legend = None
    graf_linha.x_axis.number_format = "dd-mm-yy"
    graf_linha.x_axis.majorTimeUnit = "years"
    graf_linha.y_axis.majorGridlines.graphicalProperties = GraphicalProperties(ln=LineProperties(prstDash="dash"))

    # Adiciona uma linha de tendência ao gráfico
    s1.trendline = Trendline()

    #Adiciona o gráfico à planilha
    estatisticas.add_chart(graf_linha, "L3")
    planilha.save(arquivo)

def graf_stock(base, arquivo, estilo):
    """Gera um gráfico do tipo Stock em uma planilha excel já existente. É necessário que exista uma Worksheet na planilha chamada Estatísticas.

    Args:
        base (dict): Recebe um dicionário com o nome das ações como Keys, e como valor um dataframe com os dados open-close-high-low da respectiva ação
        arquivo (str): Nome do arquivo excel onde será gerado o gráfico. Deve estar no formato ("nome.xlsx")
        estilo (openpyxl.style): deve estar no formato  estilo_titulo_carteira = NamedStyle(name = "estilo_titulo_carteira")
    """    

    # Essa função irá criar um Stock Chart para cada ativo com as informações de cada um no último ano, de 
    # semana a semana. Não há como inserir uma legenda explicativa no gráfico, mas ele deve ser lido assim:
    # As linhas determinam o preço mais alto e mais baixo que a ação atingiu em cada dia.
    # Caso a caixa esteja clara, isso indica um aumento no preço, logo a extremidade inferior da caixa
    # indica o valor de abertura, e a extremidade superior indica o valor de fechamento. Caso a caixa
    # esteja escura, houve um decrescimento, logo a extremidade superior indica o valor de abertura,
    # e a inferior indica o valor de fechamento

    # Vamos começar importando a planilha para inserir os dados nela
    planilha = load_workbook(arquivo)
    estatisticas = planilha["Estatísticas"]

    # É necessário iterar sobre o dataframe dado, para inserir os dados de cada StockChart na planilha.
    # Para iterar sobre um dataframe sem utilizar muitos recursos de Pandas, foi utilizado um código mais
    # grosseiro de For dentro de For, com um contador para separar os gráficos, um para as colunas, e um
    # para as linhas
    contador = 0
    for key in base.keys():
        contador += 1 #Este contador cresce de acordo com a quantidade de ativos que são dados
        dataframe = base[key]
        
        linha = 2
        coluna = 9 * contador #Perceba que os dados são inseridos em colunas múltiplas de 9
        estatisticas.cell(row=linha, column=coluna, value="Date")
        for item in dataframe.index:
            item = item.strftime("%Y-%m-%d")
            linha += 1
            estatisticas.cell(row=linha, column=coluna, value=item)
        
        linha = 2
        coluna += 1
        estatisticas.cell(row=linha, column=coluna, value="Open")
        for item in dataframe.Open:
            linha += 1
            estatisticas.cell(row=linha, column=coluna, value=item)

        linha = 2
        coluna += 1
        estatisticas.cell(row=linha, column=coluna, value="High")
        for item in dataframe.High:
            linha += 1
            estatisticas.cell(row=linha, column=coluna, value=item)

        linha = 2
        coluna += 1
        estatisticas.cell(row=linha, column=coluna, value="Low")
        for item in dataframe.Low:
            linha += 1
            estatisticas.cell(row=linha, column=coluna, value=item)

        linha = 2
        coluna += 1
        estatisticas.cell(row=linha, column=coluna, value="Close")
        for item in dataframe.Close:
            linha += 1
            estatisticas.cell(row=linha, column=coluna, value=item)
        aplicar_estilo_area(estatisticas,2,linha,coluna-5,coluna,estilo)

    # Aqui cria-se um gráfico para cada ativo. Para não sobrepor os gráficos, vamos diferenciá-los através 
    # de um contador. Usaremos do posicionamento dos dados ser em múltiplos de 9 para localizá-los na planilha,
    #  e dessa forma não precisaremos nos referir a cada um em específico
    grafico = 0
    for key in base.keys():
        grafico += 1
        stock = StockChart()   
        dataframe = base[key]
        
        #Primeiro selecionamos os dias e os dados na planilha
        coluna = 9 * grafico
        num_linhas = len(dataframe.index)
        dias = Reference(estatisticas, min_col=coluna, max_col=coluna, min_row=3, max_row=num_linhas)
        
        coluna_ini = coluna + 1
        coluna_fin = coluna + 4
        dados = Reference(estatisticas, min_col=coluna_ini, max_col=coluna_fin, min_row=2, max_row=num_linhas) 

        stock.add_data(dados, titles_from_data=True)
        stock.set_categories(dias)

        #Aqui configuramos as linhas High-Low e as caixas de Open-Close
        stock.hiLowLines = ChartLines()
        stock.upDownBars = UpDownBars(gapWidth=10)
        
        # Adicionando títulos e estilizando
        for serie in stock.series:
            serie.graphicalProperties.line.noFill = True

        stock.title = f"{key} Stock Chart"
        stock.legend = None
        stock.style = 3
        stock.y_axis.title = "Valor da Ação (em R$)"
        stock.x_axis.title = "Data da Informação"
        stock.x_axis.number_format = "dd-mm-yy"
        stock.x_axis.majorTimeUnit = "years"
        
        # Devido um bug do Excel, é necessário criar uma Serie dummy com valores especificos 
        # para que as linhas high-low apareçam
        pts = [NumVal(idx=i) for i in range(len(dados) - 1)]
        cache = NumData(pt=pts)
        stock.series[-1].val.numRef.numCache = cache

        # Vamos organizar os gráficos stock em duas colunas, com uma leve matemágica
        if grafico % 2 != 0:
            pos_coluna = "B"
            pos_linha = int(2 + (16 * ((grafico + 1)/2)))
        else:
            pos_coluna = "L"
            pos_linha = int(2 + (16*(grafico/2)))
        estatisticas.add_chart(stock, f"{pos_coluna}{pos_linha}")

    planilha.save(arquivo)

"""==========================================================================================="""

# Esta parte do código é apenas para teste, e demora consideravelmente para rodar!

# from cotacao import cotacao_anual

# carteira = {
#     "AVST.L": "213.1243", 
#     "ANTO.L	": "32",
#     "PETZ3.SA": "501.49", 
#     "TWDBRL=X": "1203.000000",
#     "9988.HK": "123.1", 
#     "0700.HK": "34", 
#     "ARSBRL=X": "4300.21345",
#     "INRBRL=X": "125214.3214",
#     "CHFBRL=X": "12465"
#     }
# base = cotacao_anual(carteira)

# graf_barras(carteira, "teste.xlsx")
# graf_linha(carteira, base, "teste.xlsx")
# graf_stock(base, "teste.xlsx")
 