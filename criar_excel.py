from openpyxl import Workbook, load_workbook

def criar_arquivo(nome_arquivo):
    """Cria o arquivo excel

    Args:
        nome_arquivo (string): deve estar no formato "nome_arquivo.xlsx"
    """    
    
    planilha = Workbook()
    planilha.save(nome_arquivo)


def criar_folhas(nome_arquivo):
    """Coloca as folhas nomeadas

    Args:
        nome_arquivo (string): deve estar no formato "nome_arquivo.xlsx"
    """
    
    #Abre o arquivo
    planilha = load_workbook(nome_arquivo)
    folha1 = planilha.active
    #Criar Folhas nomeadas
    folha1.title = "Carteira"
    planilha.create_sheet("Estatísticas", 1)
    #Salva a planilha
    planilha.save(nome_arquivo)

def link_folhas(nome_arquivo):
    """Cria o Link entre as duas folhas

    Args:
        nome_arquivo (string): deve estar no formato "nome_arquivo.xlsx"
    """
    
    #Abre o arquivo
    planilha = load_workbook(nome_arquivo)
    #Seleciona as folhas
    carteira = planilha["Carteira"]
    estatisticas = planilha["Estatísticas"]
    #Cria as strings dos links
    link_para_estatisticas = f"=HYPERLINK(\"[{nome_arquivo}]Estatísticas!A1\",\"Ir para Estatísticas ->\")"
    link_para_carteira = f"=HYPERLINK(\"[{nome_arquivo}]Carteira!A1\",\"Ir para Carteira ->\")"
    #Adiciona o link a celula A1 de cada folha
    carteira.merge_cells("A1:C1")
    carteira["A1"] = link_para_estatisticas
    estatisticas.merge_cells("A1:C1")
    estatisticas["A1"] = link_para_carteira
    #Salva a planilha
    planilha.save(nome_arquivo)

def aplicar_estilo_area(nome_folha, linha_inicial, linha_final, coluna_inicial, coluna_final, estilo):
    """Aplica um estilo de formatação de célula a uma range de células

    Args:
        nome_folha (openpyxl.worksheet.worksheet.Worksheet): deve estar no formato load_workbook(nome_arquivo)["nome_folha"]
        linha_inicial (int): 
        linha_final (int): 
        coluna_inicial (int): 
        coluna_final (int): 
        estilo (openpyxl.style): deve estar no formato  estilo_titulo_carteira = NamedStyle(name = "estilo_titulo_carteira")
    """    
    linhas = range(linha_inicial, linha_final + 1)
    colunas = range(coluna_inicial, coluna_final + 1)
    for linha in linhas:
        for coluna in colunas:
            nome_folha.cell(row = linha, column = coluna).style = estilo

def bloquear_planilha(nome_arquivo):
    """Bloqueia as folhas da planilha

    Args:
        nome_arquivo (string): deve estar no formato "nome_arquivo.xlsx"
    """    
    #Abre o arquivo
    planilha = load_workbook(nome_arquivo)
    #Seleciona as folhas
    carteira = planilha["Carteira"]
    estatisticas = planilha["Estatísticas"]
    #Remover gridlines
    carteira.sheet_view.showGridLines = False
    estatisticas.sheet_view.showGridLines = False
    #Remover headers
    carteira.sheet_view.showRowColHeaders = False
    estatisticas.sheet_view.showRowColHeaders = False
    carteira.protection.password = "protegida"
    estatisticas.protection.password = "protegida"
    #Salva a planilha
    planilha.save(nome_arquivo)

def criar_planilha(nome_arquivo):
    """Cria o a planilha base

    Args:
        nome_arquivo (string): deve estar no formato "nome_arquivo.xlsx"
    """
    criar_arquivo(nome_arquivo)
    criar_folhas(nome_arquivo)
    link_folhas(nome_arquivo)
