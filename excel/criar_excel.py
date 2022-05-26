from openpyxl import Workbook, load_workbook

#Cria a planilha
def criar_planilha(nome_da_planilha):
    planilha = Workbook()
    nome_arquivo = nome_da_planilha + ".xlsx"
    planilha.save(nome_arquivo)

#Coloca as folhas nomeadas
def criar_folhas(nome_arquivo):
    #Abre o arquivo
    planilha = load_workbook(nome_arquivo)
    folha1 = planilha.active
    #Criar
    folha1.title = "Carteira"
    folha2 = planilha.create_sheet("Estatísticas", 1)
    folha3 = planilha.create_sheet("YFINANCE", 2)

    planilha.save(nome_arquivo)

criar_planilha("teste")
criar_folhas("teste.xlsx")
