from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle,PatternFill, Border, Side, Alignment, Protection, Font
from cotacao import cotacao_semanal, cotacao_atual
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import qrcode

#Funções de estilo da carteira
def aplicar_estilo_area(nome_folha, linha_inicial, linha_final, coluna_inicial, coluna_final, estilo):
    """Aplica um estilo de formatação de célula a uma range de células

    Args:
        nome_folha (openpyxl.worksheet.worksheet.Worksheet): deve estar no formato load_workbook(nome_arquivo)["nome_folha"]
        linha_inicial (int): 
        linha_final (int): 
        coluna_inicial (int): 
        coluna_final (int): 
        estilo (openpyxl.style): deve estar no formato  estilo_titulo_carteira = NamedStyle(name = "estilo_titulo_carteira")
    """    
    linhas = range(linha_inicial, linha_final + 1)
    colunas = range(coluna_inicial, coluna_final + 1)
    for linha in linhas:
        for coluna in colunas:
            nome_folha.cell(row = linha, column = coluna).style = estilo

def estilizar_cabecalho(nome_folha):
    """Estiliza o cabeçalho da carteira

    Args:
        nome_folha (openpyxl.worksheet.worksheet.Worksheet): deve estar no formato load_workbook(nome_arquivo)["nome_folha"]
    """    
    #Define o estilo utilizado em todo o cabeçalho
    estilo_padrao_cabecalho = NamedStyle(name = "estilo_padrao_cabecalho")
    estilo_padrao_cabecalho.fill = PatternFill(fill_type = "solid", fgColor="00333333")
    #Aplica o estilo em todo o cabeçalho
    aplicar_estilo_area(nome_folha, 2, 8, 1, 16, estilo_padrao_cabecalho)
    #Define o estilo do título do título da carteira
    estilo_titulo_carteira = NamedStyle(name = "estilo_titulo_carteira")
    estilo_titulo_carteira.font = Font(name = "Arial Black", size = 24,color = "FFFFFF")
    estilo_titulo_carteira.fill = PatternFill(fill_type = "solid", fgColor="00333333")
    estilo_titulo_carteira.alignment = Alignment(horizontal = "center", vertical = "center")
    #Aplica o estilo somente no título da carteira
    nome_folha["B3"].style = estilo_titulo_carteira
    #Define o estilo da célula "Valor Total:"
    estilo_valor_total = NamedStyle(name = "estilo_valor_total")
    estilo_valor_total.font = Font(name = "Arial Black", size = 24)
    estilo_valor_total.alignment = Alignment(horizontal = "center", vertical = "center")
    estilo_valor_total.fill = PatternFill(fill_type="solid", fgColor="00C0C0C0")
    #Aplica o estilo somente na célula "Valor Total:"
    nome_folha["L3"].style = estilo_valor_total

def estilos_tabela_ativo():
    """Retorna um dicionário com os estilos utilizados para as tabelas por ativo
    """    
    estilos = dict()
    #Define o estilo utilizado da margem da tabela
    estilo_margem = NamedStyle(name = "estilo_margem")
    estilo_margem.fill = PatternFill(fill_type = "solid", fgColor = "00333333")
    estilos["estilo_margem"] = estilo_margem
    #Define o estilo do background
    estilo_background = NamedStyle(name = "estilo_background")
    estilo_background.fill = PatternFill(fill_type = "solid", fgColor = "00808080")
    estilos["estilo_background"] = estilo_background
    #Define o estilo dos títulos
    estilo_titulos = NamedStyle(name = "estilo_titulos")
    estilo_titulos.fill = PatternFill(fill_type = "solid", fgColor = "00C0C0C0")
    estilo_titulos.font = Font(name = "Arial Black")
    estilo_titulos.alignment = Alignment(horizontal = "center", vertical="center")
    estilos["estilo_titulos"] = estilo_titulos
    #Define o estilo dos blocos
    estilo_blocos = NamedStyle(name = "estilo_blocos")
    estilo_blocos.fill = PatternFill(fill_type = "solid", fgColor = "00FFFFFF")
    estilo_blocos.font = Font(name = "Arial Black")
    estilo_blocos.alignment = Alignment(horizontal = "center", vertical = "center")
    estilos["estilo_blocos"] = estilo_blocos
    #Define o estilo do header da tabela principal
    estilo_header = NamedStyle(name = "estilo_header")
    estilo_header.fill = PatternFill(fill_type = "solid", fgColor = "00000000")
    estilo_header.font = Font(name = "Arial Black", color = "00FFFFFF", size = 10)
    estilo_header.alignment = Alignment(horizontal = "center", vertical = "center")
    estilos["estilo_header"] = estilo_header
    #Define o estilo das linhas da tabela principal
    estilo_linhas_tabela = NamedStyle(name = "estilo_linhas_tabela")
    estilo_linhas_tabela.fill = PatternFill(fill_type = "solid", fgColor = "00FFFFFF")
    estilo_linhas_tabela.font = Font(name = "Arial", color = "00000000", size = 9)
    estilo_linhas_tabela.alignment = Alignment(horizontal = "center", vertical = "center")
    estilos["estilo_linhas_tabela"] = estilo_linhas_tabela
    
    return estilos

def estilizar_tabela_ativo(nome_folha, estilos_tabela_ativos, linha_inicial, coluna_inicial):
    """Estiliza as tabelas por ativo

    Args:
        nome_folha (openpyxl.worksheet.worksheet.Worksheet): deve estar no formato load_workbook(nome_arquivo)["nome_folha"]
        estilos_tabela_ativos (dictionary): deve estar no formato {titulo1:estilo1,titulo2:estilo2,...}
        linha_inicial (int): linha em que começa o bloco da tabela
        coluna_inicial (int): coluna em que começa o bloco
    """    
    #Estiliza a margem
    estilo_margem = estilos_tabela_ativos["estilo_margem"]
    aplicar_estilo_area(nome_folha, linha_inicial, linha_inicial+11, coluna_inicial, coluna_inicial+15,estilo_margem)
    #Estiliza o background
    estilo_background = estilos_tabela_ativos["estilo_background"]
    aplicar_estilo_area(nome_folha, linha_inicial, linha_inicial+10, coluna_inicial+1, coluna_inicial+14, estilo_background)
    #Estiliza os titulos
    estilo_titulos = estilos_tabela_ativos["estilo_titulos"]
    #Nome da ação
    nome_folha.cell(row = linha_inicial+1, column = coluna_inicial+2).style = estilo_titulos
    #Título blocos de resumo
    nome_folha.cell(row = linha_inicial+1, column = coluna_inicial+11).style = estilo_titulos
    nome_folha.cell(row = linha_inicial+4, column = coluna_inicial+11).style = estilo_titulos
    nome_folha.cell(row = linha_inicial+7, column = coluna_inicial+11).style = estilo_titulos
    #Estiliza os blocos
    estilo_blocos = estilos_tabela_ativos["estilo_blocos"]
    nome_folha.cell(row = linha_inicial+1, column = coluna_inicial+12).style = estilo_blocos
    nome_folha.cell(row = linha_inicial+4, column = coluna_inicial+12).style = estilo_blocos
    nome_folha.cell(row = linha_inicial+7, column = coluna_inicial+12).style = estilo_blocos
    #Estiliza os headers
    estilo_header = estilos_tabela_ativos["estilo_header"]
    aplicar_estilo_area(nome_folha, linha_inicial+4, linha_inicial+4, coluna_inicial+2, coluna_inicial+9, estilo_header)
    #Estiliza as linhas da tabela
    estilo_linhas_tabela = estilos_tabela_ativos["estilo_linhas_tabela"]
    aplicar_estilo_area(nome_folha,linha_inicial+5, linha_inicial+9, coluna_inicial+2, coluna_inicial+9, estilo_linhas_tabela)


def formatar_valor(nome_folha, formato, linha_inicial, linha_final, coluna_inicial, coluna_final):
    """Formata um range de células como real

    Args:
        nome_folha (openpyxl.worksheet.worksheet.Worksheet): deve estar no formato load_workbook(nome_arquivo)["nome_folha"]
        linha_inicial (int): 
        linha_final (int): 
        coluna_inicial (int): 
        coluna_final (int): 
    """    
    linhas_formatar = range(linha_inicial, linha_final+1)
    colunas_formatar = range(coluna_inicial, coluna_final+1)
    for linha in linhas_formatar:
        for coluna in colunas_formatar:
            if formato == "Real":
                nome_folha.cell(row = linha, column = coluna).number_format = "R$ #,##0.00"
            elif formato == "Data":
                nome_folha.cell(row = linha, column = coluna).number_format = "dd-mm-yyyy"
def ajustar_largura_colunas(nome_folha):
    nome_folha.column_dimensions['A'].width = 2
    nome_folha.column_dimensions['B'].width = 7
    nome_folha.column_dimensions['C'].width = 10
    nome_folha.column_dimensions['D'].width = 15
    nome_folha.column_dimensions['E'].width = 15
    nome_folha.column_dimensions['F'].width = 15
    nome_folha.column_dimensions['G'].width = 15
    nome_folha.column_dimensions['H'].width = 21
    nome_folha.column_dimensions['I'].width = 15
    nome_folha.column_dimensions['J'].width = 15
    nome_folha.column_dimensions['K'].width = 2
    nome_folha.column_dimensions['L'].width = 20
    nome_folha.column_dimensions['M'].width = 20
    nome_folha.column_dimensions['N'].width = 7.2
    nome_folha.column_dimensions['O'].width = 7.2
    nome_folha.column_dimensions['P'].width = 2

#Funções de adição dos dados
def criar_qrcode(valores_atuais,nome_qrcode):
    valor_total = 0
    for valor in valores_atuais.values():
        valor = float(valor)
        valor_total += valor
    valor_total = round(valor_total,2)
    valor_total = str(valor_total).replace('.',',')
    mensagem = str(f"O valor total da sau carteira é R$ {valor_total}")
    img = qrcode.make(mensagem)
    img.save(nome_qrcode)

def criar_cabecalho(nome_folha):
    """Cria o cabeçalho da planilha

    Args:
        nome_folha (openpyxl.worksheet.worksheet.Worksheet): deve estar no formato load_workbook(nome_arquivo)["nome_folha"]
    """    
    #Cria o título da carteira
    nome_folha["B3"]  = "Carteira de Ativos"
    nome_folha.merge_cells("B3:K7")
    #Criar espaço para QRCODE
    nome_folha["L3"] = "Valor Total:"
    nome_folha.merge_cells("L3:M7")
    nome_folha["N3"] =  "QRCODE"
    nome_folha.merge_cells("N3:O7")
    estilizar_cabecalho(nome_folha)

def criar_tabela_ativo(nome_folha, ativo, data_frame, linha_inicial, coluna_inicial):
    """Cria na planilha a tabela por ativo

    Args:
        nome_folha (openpyxl.worksheet.worksheet.Worksheet): deve estar no formato load_workbook(nome_arquivo)["nome_folha"]
        ativo (string): nome do ativo que será o título
        data_frame (dataframe): dataframe que se tornará a tabela
        linha_inicial (int): linha em que começa o bloco da tabela
        coluna_inicial (int): coluna em que começa o bloco
    """    
    #Cria o título do ativo
    linha_titulo = linha_inicial + 1
    nome_folha.cell(row = linha_titulo,column = coluna_inicial, value = ativo)
    nome_folha.merge_cells(start_row = linha_titulo, start_column = coluna_inicial, end_row = linha_titulo+1, end_column = coluna_inicial +  7)
    #Define a linha inicial da tabela
    linha_tabela = linha_inicial + 4 
    #Transforma o dataframe em uma lista de linhas
    linhas_data_frame = list(dataframe_to_rows(data_frame,index=True,header=True))
    """
    O rótulo date estava uma linha abaixo como frozen list, foi necessário converte-lo e adicioná-lo a linha correta
    """
    lista_rotulo_date = list(linhas_data_frame[1])
    rotulo_date =  str(lista_rotulo_date[0])
    linhas_data_frame[0][0] = rotulo_date
    del linhas_data_frame[1]
    #Acrescenta a tabela dataframe
    for linha in linhas_data_frame:
        #Pega cada lista e coloca em uma linha, com cada item em uma coluna
        coluna_item = coluna_inicial
        for item_coluna in linha:
            nome_folha.cell(row = linha_tabela, column = coluna_item, value = item_coluna)
            coluna_item += 1
        linha_tabela += 1
    #Formata os valores ao tipo correto
    formatar_valor(nome_folha, "Real", linha_inicial + 5, linha_inicial + 9, coluna_inicial + 1, coluna_inicial + 6)
    formatar_valor(nome_folha, "Data", linha_inicial + 5, linha_inicial + 9, coluna_inicial, coluna_inicial)

def criar_blocos(linha_inicial, coluna_inicial, titulo, valor, nome_folha):
    """Cria um dois blocos no excel, com tammanho 2x2 um ao lado do outro

    Args:
        linha_inicial (int): linha em que começa os blocos da tabela
        coluna_inicial (int): coluna em que começa o bloco
        titulo (string): Impresso no bloco da esquerda
        valor (Any): Impresso no bloco da direita
        nome_folha (openpyxl.worksheet.worksheet.Worksheet): deve estar no formato load_workbook(nome_arquivo)["nome_folha"]
    """    
    nome_folha.cell(row = linha_inicial,column = coluna_inicial, value = titulo)
    nome_folha.merge_cells(start_row = linha_inicial, start_column = coluna_inicial, end_row = linha_inicial+1, end_column = coluna_inicial)
    nome_folha.cell(row = linha_inicial,column = coluna_inicial+1, value = valor)
    nome_folha.merge_cells(start_row = linha_inicial, start_column = coluna_inicial+1, end_row = linha_inicial+1, end_column = coluna_inicial +  2)

def criar_resumo(nome_folha, valor_ultima_cotacao, quantidade, linha_inicial, coluna_inicial):
    """Cria os blocos de resumo do ativo

    Args:
        nome_folha (openpyxl.worksheet.worksheet.Worksheet): deve estar no formato load_workbook(nome_arquivo)["nome_folha"]
        valor_ultima_cotacao (float): deve estar no formato 9999.99
        quantidade (float): deve estar no formato 9999.99
        linha_inicial (int): linha em que começa o bloco de resumos
        coluna_inicial (int): coluna em que começa o bloco de resumos
    """    
    #Cria os blocos do Total Anual
    linha_total = linha_inicial +1
    valor_total = valor_ultima_cotacao * quantidade
    criar_blocos(linha_total, coluna_inicial, "Total Atual", valor_total, nome_folha)
    formatar_valor(nome_folha, "Real", linha_total,linha_total, coluna_inicial + 1, coluna_inicial + 1)
    #Cria os blocos de Qtd. Ativos
    linha_quantidade = linha_inicial + 4
    criar_blocos(linha_quantidade, coluna_inicial, "Qtd. Ativos", quantidade, nome_folha)
    #Cria o bloco de Última Cotação
    linha_ultima_cotacao = linha_inicial + 7
    criar_blocos(linha_ultima_cotacao, coluna_inicial, "Última Cotação", valor_ultima_cotacao, nome_folha)
    formatar_valor(nome_folha, "Real", linha_ultima_cotacao, linha_ultima_cotacao, coluna_inicial + 1, coluna_inicial + 1)

def criar_corpo_carteira(nome_folha, dicionario_ativos):
    """Cria todas as tabelas por cada ativo

    Args:
        nome_folha (openpyxl.worksheet.worksheet.Worksheet): deve estar no formato load_workbook(nome_arquivo)["nome_folha"]
        dicionario_ativos (dictionary): deve estar no formato {ativo1:qtd1,ativo2:qtd2,...}
    """    
    #Puxa o dicionario de ativos e seus dataframes
    cotacao = cotacao_semanal(dicionario_ativos)
    #Puxa os valores das ultimas cotações
    valores_atuais = cotacao_atual(cotacao)
    #Define a linha que termina o cabeçalho
    linha_inicial = 9
    #Define os estilos para estilizar as tabelas
    estilos = estilos_tabela_ativo()
    #Para cada ativo cria o bloco do ativo
    for ativo, data_frame in cotacao.items():
        estilizar_tabela_ativo(nome_folha, estilos, linha_inicial, 1)
        criar_tabela_ativo(nome_folha, ativo, data_frame, linha_inicial, 3)
        valor_atual = valores_atuais[ativo]
        quantidade_ativo = float(dicionario_ativos[ativo])
        criar_resumo(nome_folha, valor_atual, quantidade_ativo, linha_inicial, 12)
        #Define o começo do próximo bloco
        linha_inicial += 12

#Função final
def carteira(nome_arquivo, dicionario_ativos):
    """Realiza todas as modificações na folha carteira

    Args:
        nome_arquivo (string): deve estar no formato("nome_arquivo.xlsx")
        dicionario_ativos (dictionary): deve estar no formato {ativo1:qtd1,ativo2:qtd2,...}
    """    
    #Abre a planilha
    planilha = load_workbook(nome_arquivo)
    #Seleciona a folha da carteira
    folha_carteira = planilha["Carteira"]
    criar_cabecalho(folha_carteira)
    criar_corpo_carteira(folha_carteira, dicionario_ativos)
    ajustar_largura_colunas(folha_carteira)
    #Salva o arquivo
    planilha.save(nome_arquivo)

dic = {'PETR4.SA': '240', 'B3SA3.SA': '120', 'HAPV3.SA': '300', 'OIBR3.SA':'78','BRL=X':'3187.76','JPYBRL=X':'120987.09','EURBRL=X':'2490.87'}
#from criar_excel import criar_planilha
#criar_planilha("Teste1.xlsx")
#carteira("Teste1.xlsx",dic)
print(criar_qrcode(dic,"qrcode_teste1.png"))